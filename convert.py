from openpyxl import load_workbook

xlsx_file = 'TBM Organization Parameters.xlsx'

# Load the workbook
workbook = load_workbook(xlsx_file)

ignrshts = ['WiFi_Languages','Validation','Data Valid']

spclsht = ['Wifi_Materials','Wifi_Downloads']

# Iterate through all the sheet names
for sheet_name in workbook.sheetnames:
    # Access each sheet by name
    sheet = workbook[sheet_name]
    if sheet.title in spclsht:
        if sheet.title == 'Wifi_Materials':
            rctr = 1
            with open('diff/WiFi_Materials.txt', 'w') as file:
                for row in sheet.iter_rows(values_only=True):
                    if rctr > 1:
                        file.write('Material: ' + row[0]+'\n')
                        file.write('Orgs: ' + '\t' + row[1]+'\n')
                        file.write('Langs: ' + '\t' + row[2]+'\n')
                        file.write('\n')
                    rctr+=1
        elif sheet.title == 'Wifi_Downloads':
            rctr = 1
            with open('diff/WiFi_Downloads.txt', 'w') as file:
                for row in sheet.iter_rows(values_only=True):
                    if rctr > 1 and row[0]:
                        file.write('Name: ' + row[0]+'\n')
                        file.write('Cat: ' + '\t' + row[1]+'\n')
                        file.write('Langs: ' + '\t' + row[2]+'\n')
                        file.write('\n')
                    else:
                        print(row)
                    rctr+=1

    elif sheet.title not in ignrshts:
        rctr = 1
        with open(f'diff/{sheet.title}.txt', 'w') as file:
            vills = []
            zons = []
            tribs = []
            langs = []
            mats = []
            for row in sheet.iter_rows(values_only=True):
                if rctr == 1:
                    file.write('Org Name: ' + row[1]+'\n')
                    if len(row) > 6:
                        file.write('Country: ' + row[6]+'\n')
                if rctr > 2:
                    if row[0]:
                        vills.append(row[0])
                    if row[1]:
                        zons.append(row[1])
                    if row[2]:
                        tribs.append(row[2])
                    if row[3]:
                        langs.append(row[3])
                    if row[4]:
                        mats.append(row[4])
                rctr+=1
            # Write out Villages
            file.write('\nVillages:\n')
            for vil in vills:
                file.write(f'\t{vil}\n')
            # Write out Zones
            file.write('\nZones:\n')
            for zon in zons:
                file.write(f'\t{zon}\n')
            # Write out Tribes
            file.write('\nTribes:\n')
            for trib in tribs:
                file.write(f'\t{trib}\n')
            # Write out Languages
            file.write('\nLanguages:\n')
            for lang in langs:
                file.write(f'\t{lang}\n')
            # Write out Materials
            file.write('\nMaterials:\n')
            for mat in mats:
                file.write(f'\t{mat}\n')
    
    else:
        print(f'{sheet.title} Ignored\n')
    # Iterate through rows in the sheet (optional)
    #for row in sheet.iter_rows(values_only=True):
    #    print(row)  # Print each row in the sheet

print("Conversion completed successfully.")